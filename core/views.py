# core/views.py
import json
import io
import openpyxl

from django.contrib import messages
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.exceptions import ObjectDoesNotExist
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import JsonResponse, HttpResponse, HttpResponseRedirect
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from openpyxl.utils import get_column_letter

from .models import (
    empleado,
    liquidacion,
    jornada,
    turno_has_jornada,
    ZonaTrabajo,
    contrato,
    turno,
    cargo,
    Foto,
)
from .forms import (
    EmpleadoZonaForm,
    ZonaTrabajoForm,
    ContratoForm,
    CargoForm,
    PerfilForm,
    LoginForm,
)

User = get_user_model()



# ---------------------------
# Helpers
# ---------------------------

@login_required
@user_passes_test(lambda u: u.is_staff, login_url='/')
def horario_admin_page(request):
    # Mantiene tu comportamiento anterior: redirigir al listado de horarios
    return redirect('horario_jornada')


def _empleado_de_usuario(user):
    return empleado.objects.select_related('user').get(user=user)


def _contrato_actual(emp: empleado):
    """
    Trae el contrato más reciente del empleado con relaciones para
    evitar 'object (...)' en plantillas.
    """
    return (
        contrato.objects
        .select_related(
            'empleado',
            'cargo',
            'departamento',
            'turno_has_jornada',
            'turno_has_jornada__turno',
            'turno_has_jornada__jornada',
        )
        .filter(empleado=emp)
        .order_by('-fecha_inicio', '-id')
        .first()
    )


def _rol_de(user):
    if user.is_superuser or user.is_staff or user.groups.filter(name__iexact='Admin').exists():
        return 'admin'
    return 'empleado'


def is_admin(u):
    return u.is_authenticated and u.is_staff


# ---------------------------
# Auth / sesión
# ---------------------------

def login_page(request):
    if request.user.is_authenticated:
        return redirect('dash_admin' if _rol_de(request.user) == 'admin' else 'dash_empleado')
    return render(request, 'rrhh/login.html')


@csrf_exempt
def login_json(request):
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'msg': 'Método no permitido'}, status=405)

    try:
        data = json.loads(request.body.decode('utf-8'))
    except Exception:
        data = request.POST

    email = (data.get('email') or '').strip().lower()
    password = data.get('password') or ''

    if not email or not password:
        return JsonResponse({'ok': False, 'msg': 'Faltan credenciales'}, status=400)

    try:
        u = User.objects.get(email__iexact=email)
        username = u.get_username()
    except User.DoesNotExist:
        return JsonResponse({'ok': False, 'msg': 'Usuario no encontrado o inactivo'}, status=401)

    user = authenticate(request, username=username, password=password)
    if not user or not user.is_active:
        return JsonResponse({'ok': False, 'msg': 'Credenciales inválidas'}, status=401)

    login(request, user)
    rol = _rol_de(user)
    return JsonResponse({'ok': True, 'user': {'id': user.id, 'email': user.email, 'rol': rol}})


def logout_view(request):
    logout(request)
    return redirect('login_page')


@login_required
def me(request):
    return JsonResponse({'email': request.user.email, 'rol': _rol_de(request.user)})


# ---------------------------
# Dashboards
# ---------------------------

@login_required
def dashboard_empleado(request):
    if _rol_de(request.user) != 'empleado':
        return redirect('dash_admin')

    emp = _empleado_de_usuario(request.user)
    cto = _contrato_actual(emp)

    liqs = (
        liquidacion.objects
        .filter(contrato__empleado=emp)
        .order_by('-periodo')[:2]
    )

    horarios = []
    thj = None

    if cto and getattr(cto, "turno_has_jornada_id", None):
        try:
            thj = cto.turno_has_jornada
        except ObjectDoesNotExist:
            thj = None

    if thj and getattr(thj, "turno", None) and getattr(thj, "jornada", None):
        horarios = [{
            "dia": "-",
            "entrada": thj.turno.hora_entrada,
            "salida": thj.turno.hora_salida,
            "descanso": "-",
            "observacion": thj.jornada.nombre,
        }]
    else:
        horarios = [{
            "dia": "-",
            "entrada": "-",
            "salida": "-",
            "descanso": "-",
            "observacion": "-",
        }]

    return render(request, 'rrhh/dashboard.html', {
        "emp": emp,
        "contrato": cto,
        "liquidaciones": liqs,
        "horarios": horarios,
    })


@login_required
def dashboard_admin(request):
    if _rol_de(request.user) != 'admin':
        return redirect('dash_empleado')

    try:
        admin_emp = empleado.objects.select_related('user', 'zona_trabajo').get(user=request.user)
    except empleado.DoesNotExist:
        admin_emp = None

    admin_liqs = []
    if admin_emp:
        admin_liqs = (
            liquidacion.objects
            .filter(contrato__empleado=admin_emp)
            .order_by('-periodo')[:2]
        )

    q = request.GET.get('q', '').strip()
    empleados_qs = empleado.objects.select_related('user', 'zona_trabajo').order_by('id')
    if q:
        empleados_qs = empleados_qs.filter(
            Q(user__first_name__icontains=q) |
            Q(user__last_name__icontains=q) |
            Q(run__icontains=q) |
            Q(zona_trabajo__nombre__icontains=q)
        )

    empleados_data = []
    for e in empleados_qs:
        cto = _contrato_actual(e)
        cargo_nombre = cto.cargo.nombre if (cto and cto.cargo) else "—"

        horario_str = "-"
        if cto and getattr(cto, "turno_has_jornada_id", None):
            try:
                thj = cto.turno_has_jornada
                trn = getattr(thj, "turno", None)
                jor = getattr(thj, "jornada", None)
                ent = getattr(trn, "hora_entrada", None)
                sal = getattr(trn, "hora_salida", None)
                nom = getattr(jor, "nombre", None)

                if ent and sal and nom:
                    ent_txt = ent.strftime("%H:%M") if hasattr(ent, "strftime") else str(ent)
                    sal_txt = sal.strftime("%H:%M") if hasattr(sal, "strftime") else str(sal)
                    horario_str = f"{ent_txt}-{sal_txt} ({nom})"
            except (ObjectDoesNotExist, AttributeError):
                horario_str = "-"

        empleados_data.append({
            "id": e.id,
            "nombre": e.user.get_full_name() or e.user.username,
            "cargo": cargo_nombre,
            "zona": e.zona_trabajo.nombre if e.zona_trabajo_id else "—",
            "horario": horario_str,
        })

    return render(request, 'rrhh/dashboard_admin.html', {
        "empleados": empleados_data,
        "q": q,
        "admin_emp": admin_emp,
        "admin_liqs": admin_liqs,
    })


# ---------------------------
# Horarios (empleado y admin)
# ---------------------------

@login_required
def horario_jornada_page(request):
    return redirect('horario_jornada')


@login_required
def horario_page(request):
    if _rol_de(request.user) != 'empleado':
        return redirect('dash_admin')

    emp = _empleado_de_usuario(request.user)
    cto = _contrato_actual(emp)

    horarios = []
    if cto and cto.turno_has_jornada_id:
        thj = cto.turno_has_jornada
        if thj and thj.turno and thj.jornada:
            horarios = [{
                "dia": "-",
                "entrada": thj.turno.hora_entrada,
                "salida": thj.turno.hora_salida,
                "descanso": "-",
                "zona": "-",
                "observacion": thj.jornada.nombre,
            }]

    return render(request, 'rrhh/horario.html', {"emp": emp, "horarios": horarios})


# ---------------------------
# Liquidaciones (empleado)
# ---------------------------

@login_required
def liquidacion_page(request):
    if _rol_de(request.user) != 'empleado':
        return redirect('dash_admin')

    emp = _empleado_de_usuario(request.user)
    qs = liquidacion.objects.filter(contrato__empleado=emp).order_by('-periodo')

    last = request.session.get('liq_last_page', 1)
    page_number = request.GET.get('page') or last
    paginator = Paginator(qs, 10)
    page_obj = paginator.get_page(page_number)

    request.session['liq_last_page'] = page_obj.number

    return render(request, 'rrhh/liquidacion.html', {
        "emp": emp,
        "page_obj": page_obj
    })


@login_required
def liquidaciones_list(request):
    """
    Lista las liquidaciones del empleado asociado al usuario logueado.
    """
    emp = None
    try:
        emp = empleado.objects.get(user=request.user)
    except empleado.DoesNotExist:
        emp = None

    qs = liquidacion.objects.none()
    if emp:
        qs = (
            liquidacion.objects
            .filter(contrato__empleado=emp)
            .select_related('contrato')
            .order_by('-periodo')
        )

    paginator = Paginator(qs, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    back_name = 'dash_admin' if _rol_de(request.user) == 'admin' else 'dash_empleado'

    context = {
        'page_obj': page_obj,
        'liqs': page_obj.object_list,
        'back_name': back_name,
    }
    return render(request, 'rrhh/liquidaciones_list.html', context)


# ---------------------------
# Contrato del empleado
# ---------------------------

@login_required
def contrato_empleado_page(request):
    """
    Muestra el contrato del usuario autenticado.
    """
    if _rol_de(request.user) != 'empleado':
        return redirect('dash_admin')

    emp = _empleado_de_usuario(request.user)
    cto = _contrato_actual(emp)

    return render(request, 'rrhh/contrato.html', {
        "emp": emp,
        "contrato": cto,
    })


# ---------------------------
# Módulo Contratos (admin)
# ---------------------------

@login_required
@user_passes_test(is_admin, login_url='/')
def contratos_admin_page(request):
    contratos_qs = (
        contrato.objects
        .select_related('empleado__user', 'departamento', 'cargo')
        .order_by('-fecha_inicio', '-id')
    )
    return render(request, 'rrhh/contratos_admin.html', {'contratos': contratos_qs})


@login_required
@user_passes_test(is_admin, login_url='/')
def contrato_create(request):
    if request.method == 'POST':
        form = ContratoForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Contrato creado correctamente.')
            return redirect('contratos_admin')
    else:
        form = ContratoForm()
    return render(request, 'rrhh/contrato_form.html', {'form': form, 'titulo': 'Nuevo contrato'})


@login_required
@user_passes_test(is_admin, login_url='/')
def contrato_edit(request, pk):
    obj = get_object_or_404(contrato, pk=pk)
    if request.method == 'POST':
        form = ContratoForm(request.POST, request.FILES, instance=obj)
        if form.is_valid():
            form.save()
            messages.success(request, 'Contrato actualizado.')
            return redirect('contratos_admin')
    else:
        form = ContratoForm(instance=obj)
    return render(request, 'rrhh/contrato_form.html', {'form': form, 'titulo': 'Editar contrato'})


@login_required
@user_passes_test(is_admin, login_url='/')
def contrato_delete(request, pk):
    obj = get_object_or_404(contrato, pk=pk)
    if request.method == 'POST':
        obj.delete()
        messages.success(request, 'Contrato eliminado.')
        return redirect('contratos_admin')
    return render(request, 'rrhh/contrato_confirm_delete.html', {'obj': obj})


# ---------------------------
# CRUD Zonas de Trabajo (admin)
# ---------------------------

@login_required
@user_passes_test(is_admin, login_url='/')
def zonas_list(request):
    q = request.GET.get("q", "").strip()

    zonas = ZonaTrabajo.objects.all().order_by("nombre")
    if q:
        zonas = zonas.filter(
            Q(nombre__icontains=q) |
            Q(area__icontains=q) |
            Q(ubicacion__icontains=q) |
            Q(supervisor__icontains=q)
        )

    paginator = Paginator(zonas, 5)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(request, "rrhh/zonas_list.html", {
        "q": q,
        "page_obj": page_obj,
    })


@login_required
@user_passes_test(is_admin, login_url='/')
def zona_create(request):
    if request.method == "POST":
        form = ZonaTrabajoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Zona creada correctamente.")
            return redirect("zonas_list")
    else:
        form = ZonaTrabajoForm()
    return render(request, "rrhh/zona_form.html", {"form": form, "titulo": "Nueva zona"})


@login_required
@user_passes_test(is_admin, login_url='/')
def zona_edit(request, pk):
    z = get_object_or_404(ZonaTrabajo, pk=pk)
    if request.method == "POST":
        form = ZonaTrabajoForm(request.POST, instance=z)
        if form.is_valid():
            form.save()
            messages.success(request, "Zona actualizada.")
            return redirect("zonas_list")
    else:
        form = ZonaTrabajoForm(instance=z)
    return render(request, "rrhh/zona_form.html", {"form": form, "titulo": "Editar zona"})


@login_required
@user_passes_test(is_admin, login_url='/')
def zona_delete(request, pk):
    z = get_object_or_404(ZonaTrabajo, pk=pk)
    if request.method == "POST":
        z.delete()
        messages.success(request, "Zona eliminada.")
        return redirect("zonas_list")
    return render(request, "rrhh/zona_confirm_delete.html", {"obj": z})


# ---------------------------
# Asignar/Cambiar zona a Empleado (admin)
# ---------------------------

@login_required
@user_passes_test(is_admin, login_url='/')
def empleado_zonas_list(request):
    q = request.GET.get("q", "").strip()
    emps = empleado.objects.select_related("user", "zona_trabajo").order_by("id")
    if q:
        emps = emps.filter(
            Q(user__first_name__icontains=q) |
            Q(user__last_name__icontains=q) |
            Q(run__icontains=q) |
            Q(zona_trabajo__nombre__icontains=q)
        )
    return render(request, "rrhh/empleado_zonas_list.html", {"empleados": emps, "q": q})


@login_required
@user_passes_test(is_admin, login_url='/')
def empleado_zona_edit(request, pk):
    emp = get_object_or_404(empleado.objects.select_related("user", "zona_trabajo"), pk=pk)
    if request.method == "POST":
        form = EmpleadoZonaForm(request.POST, instance=emp)
        if form.is_valid():
            form.save()
            messages.success(request, "Zona asignada/actualizada para el empleado.")
            return redirect("empleado_zonas_list")
    else:
        form = EmpleadoZonaForm(instance=emp)

    return render(
        request,
        "rrhh/empleado_zona_form.html",
        {"form": form, "empleado": emp, "titulo": f"Zona de {emp.user.get_full_name() or emp.user.username}"}
    )


# ---------------------------
# CRUD HORARIOS (admin)
# ---------------------------

@login_required
@user_passes_test(is_admin, login_url='/')
def horario_jornada_list(request):
    horarios = turno_has_jornada.objects.select_related('turno', 'jornada').all()
    return render(request, 'rrhh/horario_jornada.html', {"horarios": horarios})


@login_required
@user_passes_test(is_admin, login_url='/')
def horario_jornada_create(request):
    if request.method == "POST":
        nombre = request.POST.get("nombre")
        inicio = request.POST.get("hora_inicio")
        fin = request.POST.get("hora_fin")

        j = jornada.objects.create(nombre=nombre, horas_semanales=40)
        t = turno.objects.create(hora_entrada=inicio, hora_salida=fin)
        turno_has_jornada.objects.create(turno=t, jornada=j)

        messages.success(request, "Horario creado correctamente.")
        return redirect('horario_jornada')

    return redirect('horario_jornada')


@login_required
@user_passes_test(is_admin, login_url='/')
def horario_jornada_update(request, pk):
    h = get_object_or_404(turno_has_jornada, pk=pk)

    if request.method == "POST":
        h.jornada.nombre = request.POST.get("nombre")
        h.turno.hora_entrada = request.POST.get("hora_inicio")
        h.turno.hora_salida = request.POST.get("hora_fin")

        h.jornada.save()
        h.turno.save()

        messages.success(request, "Horario actualizado.")
        return redirect('horario_jornada')

    return redirect('horario_jornada')


@login_required
@user_passes_test(is_admin, login_url='/')
def horario_jornada_delete(request, pk):
    h = get_object_or_404(turno_has_jornada, pk=pk)
    h.turno.delete()
    h.jornada.delete()
    h.delete()
    messages.success(request, "Horario eliminado.")
    return redirect('horario_jornada')


# ---------------------------
# Horario simple (admin UI de apoyo)
# ---------------------------

@login_required
def horario_jornada(request):
    horarios = turno_has_jornada.objects.select_related('turno', 'jornada').all()
    return render(request, 'rrhh/horario_jornada.html', {'horarios': horarios})


@login_required
def horario_create(request):
    turnos = turno.objects.all()
    jornadas = jornada.objects.all()

    if request.method == 'POST':
        t = request.POST.get('turno')
        j = request.POST.get('jornada')
        turno_sel = turno.objects.get(pk=t)
        jornada_sel = jornada.objects.get(pk=j)

        turno_has_jornada.objects.create(turno=turno_sel, jornada=jornada_sel)

        messages.success(request, "Horario creado correctamente ✅")
        return redirect('horario_jornada')

    return render(request, 'rrhh/horario_form.html', {'turnos': turnos, 'jornadas': jornadas})


@login_required
def horario_update(request, pk):
    horario = get_object_or_404(turno_has_jornada, id=pk)
    turnos = turno.objects.all()
    jornadas = jornada.objects.all()

    if request.method == "POST":
        horario.turno = get_object_or_404(turno, id=request.POST.get('turno'))
        horario.jornada = get_object_or_404(jornada, id=request.POST.get('jornada'))
        horario.save()
        return redirect('horario_jornada')

    return render(request, 'rrhh/horario_form.html', {
        'horario': horario,
        'turnos': turnos,
        'jornadas': jornadas,
        'accion': 'Editar'
    })


@login_required
def horario_delete(request, pk):
    horario = get_object_or_404(turno_has_jornada, id=pk)
    horario.delete()
    return redirect('horario_jornada')


# ---------------------------
# CRUD Cargos
# ---------------------------

@login_required
def gestion_cargos(request):
    q = request.GET.get("q", "").strip()
    cargos_qs = cargo.objects.all().order_by("id")
    if q:
        cargos_qs = cargos_qs.filter(Q(nombre__icontains=q) | Q(description__icontains=q))
    return render(request, "rrhh/crud_cargo.html", {"cargos": cargos_qs, "q": q})


@login_required
def cargo_create(request):
    if request.method == "POST":
        form = CargoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Cargo creado correctamente.")
            return redirect("crud_cargo")
    else:
        form = CargoForm()
    return render(request, "rrhh/cargo_form.html", {"form": form, "title": "Nuevo Cargo"})


@login_required
def cargo_edit(request, pk):
    obj = get_object_or_404(cargo, pk=pk)
    if request.method == "POST":
        form = CargoForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            messages.success(request, "Cargo actualizado.")
            return redirect("crud_cargo")
    else:
        form = CargoForm(instance=obj)
    return render(request, "rrhh/cargo_form.html", {"form": form, "title": "Editar Cargo"})


@login_required
def cargo_delete(request, pk):
    obj = get_object_or_404(cargo, pk=pk)
    if request.method == "POST":
        obj.delete()
        messages.success(request, "Cargo eliminado.")
        return redirect("crud_cargo")
    return render(request, "rrhh/cargo_confirm_delete.html", {"obj": obj})


@login_required
def empleado_cargo_edit(request, pk):
    emp = get_object_or_404(empleado, pk=pk)

    if request.method == "POST":
        id_cargo = request.POST.get("cargo")
        if id_cargo:
            emp.cargo_id = id_cargo
            emp.save()
        return redirect('dash_admin')

    cargos_qs = cargo.objects.all()
    return render(request, "rrhh/empleado_cargo_edit.html", {"emp": emp, "cargos": cargos_qs})


# ---------------------------
# Perfil de usuario
# ---------------------------

@login_required
def perfil_view(request):
    return render(request, "usuarios/perfil.html")


@login_required
def perfil_edit(request):
    if request.method == "POST":
        form = PerfilForm(request.POST, request.FILES, instance=request.user)
        if form.is_valid():
            form.save()
            return redirect("perfil")
    else:
        form = PerfilForm(instance=request.user)
    return render(request, "usuarios/perfil_edit.html", {"form": form})


# ---------------------------
# Exportar liquidaciones a Excel
# ---------------------------

def export_liquidacion_excel(request, pk):
    liq = get_object_or_404(liquidacion, pk=pk)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Liquidación"

    headers = ["Mes", "Año", "Monto líquido", "Estado"]
    estado = liq.get_estado_display() if hasattr(liq, "get_estado_display") else liq.estado
    data = [liq.mes, liq.anio, liq.monto_liquido, estado]

    ws.append(headers)
    ws.append(data)

    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    filename = f'liquidacion_{liq.anio}_{liq.mes}.xlsx'
    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(resp)
    return resp


def export_liquidaciones_excel(request):
    qs = liquidacion.objects.all().order_by("-anio", "-mes")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Liquidaciones"

    headers = ["#", "Mes", "Año", "Monto líquido", "Estado"]
    ws.append(headers)

    for i, l in enumerate(qs, start=1):
        estado = l.get_estado_display() if hasattr(l, "get_estado_display") else l.estado
        ws.append([i, l.mes, l.anio, float(l.monto_liquido), estado])

    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = 'attachment; filename="mis_liquidaciones.xlsx"'
    wb.save(resp)
    return resp


# ---------------------------
# API crear empleado (dashboard admin)
# ---------------------------

@require_POST
def empleado_crear(request):
    data = json.loads(request.body)

    user = User.objects.create_user(
        username=data["email"].split("@")[0],
        email=data["email"],
        password="temp123456",
        first_name=data["first_name"],
        last_name=data["last_name"],
    )

    emp = empleado.objects.create(
        user=user,
        run=data["run"],
        fono=data.get("fono") or None,
        nacionalidad=data.get("nacionalidad") or "",
        cargo_id=data.get("cargo_id") or None,
        zona_trabajo_id=data.get("zona_id") or None,
    )

    return JsonResponse({"ok": True, "id": emp.id})



def login_view(request):
    if request.user.is_authenticated:
        return redirect('galeria')

    form = LoginForm(request, data=request.POST or None)

    if request.method == 'POST' and form.is_valid():
        login(request, form.get_user())
        return redirect('galeria')

    return render(request, 'rrhh/login.html', {'form': form})


@login_required(login_url='login')
def galeria_view(request):
    fotos = Foto.objects.filter(activa=True)
    return render(request, 'rrhh/galeria.html', {'fotos': fotos})


@login_required(login_url='login')
def logout_view(request):
    logout(request)
    return redirect('login')
